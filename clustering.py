import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
import umap
import matplotlib.pyplot as plt
from adjustText import adjust_text
import spacy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# ================================
# Modelli di linguaggio
# ================================
nlp = spacy.load("it_core_news_sm")
model = SentenceTransformer("sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2")

# ================================
# Funzioni di supporto
# ================================
def leggi_corpora(percorso_file):
    """Legge un file di testo e restituisce il contenuto come stringa."""
    with open(percorso_file, 'r', encoding='utf-8') as f:
        return f.read()


def salva_excel_auto_width(df, nome_file):
    """Salva un DataFrame in formato Excel e adatta automaticamente la larghezza delle colonne."""
    df.to_excel(nome_file, index=False)
    wb = load_workbook(nome_file)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(nome_file)
    print(f"File salvato: {nome_file}")


def parole_rappresentative(embeddings, df, cluster_col="CLUSTER", top_n=5):
    """Seleziona le frasi o parole più vicine al centroide di ciascun cluster."""
    rappresentative_idx = []
    rappresentative_dict = {}

    for cluster_id in df[cluster_col].unique():
        idx = df[df[cluster_col] == cluster_id].index
        cluster_embeddings = embeddings[idx]
        centroide = cluster_embeddings.mean(axis=0)
        dist = np.linalg.norm(cluster_embeddings - centroide, axis=1)
        top_idx = idx[np.argsort(dist)[:top_n]]
        rappresentative_idx.extend(top_idx)
        rappresentative_dict[cluster_id] = df.loc[top_idx, "CONTENUTO"].tolist()

    return df.loc[rappresentative_idx], embeddings[rappresentative_idx], rappresentative_dict


# ================================
# Analisi linguistica
# ================================
def annotazione(testo):
    """Elabora un testo restituendo frasi, token, POS e lemmi."""
    doc = nlp(testo)
    frasi = list(doc.sents)
    tokens = [token.text for token in doc]
    pos_tags = [(token.text, token.pos_) for token in doc]
    lemmi = [token.lemma_ for token in doc if not token.is_punct]
    return {
        'doc': doc,
        'frasi': frasi,
        'tokens': tokens,
        'pos_tags': pos_tags,
        'lemmi': lemmi
    }


def distribuzione_pos(pos_tags):
    """Conta le occorrenze principali per tipo di parte del discorso."""
    distribuzione = {'Verbi': 0, 'Aggettivi': 0, 'Nomi': 0, 'Pronomi': 0}
    for _, pos in pos_tags:
        if pos == 'VERB':
            distribuzione['Verbi'] += 1
        elif pos == 'ADJ':
            distribuzione['Aggettivi'] += 1
        elif pos in ('NOUN', 'PROPN'):
            distribuzione['Nomi'] += 1
        elif pos == 'PRON':
            distribuzione['Pronomi'] += 1
    return distribuzione


def conta_entita_nominate(doc):
    """Conta le entità nominate nel documento."""
    return len(list(doc.ents))


def analisi_linguistica(testi, nomi_corpus=None):
    """Esegue un'analisi linguistica di base su più testi."""
    risultati = []
    if nomi_corpus is None:
        nomi_corpus = [f"Corpus_{i+1}" for i in range(len(testi))]

    for testo, nome in zip(testi, nomi_corpus):
        dati = annotazione(testo)
        lunghezza_media = sum(len(frase) for frase in dati['frasi']) / len(dati['frasi']) if dati['frasi'] else 0
        pos_dist = distribuzione_pos(dati['pos_tags'])
        risultati.append({
            "Corpus": nome,
            "Frasi": len(dati['frasi']),
            "Token": len(dati['tokens']),
            "Lunghezza_media_frasi": round(lunghezza_media, 2),
            "Lemmi_distinti": len(set(dati['lemmi'])),
            "Verbi": pos_dist['Verbi'],
            "Aggettivi": pos_dist['Aggettivi'],
            "Nomi": pos_dist['Nomi'],
            "Pronomi": pos_dist['Pronomi'],
            "Entità_nominate": conta_entita_nominate(dati['doc'])
        })

    # Analisi complessiva del corpus unito
    testo_completo = "\n".join(testi)
    dati_completo = annotazione(testo_completo)
    lunghezza_media = sum(len(frase) for frase in dati_completo['frasi']) / len(dati_completo['frasi']) if dati_completo['frasi'] else 0
    pos_dist = distribuzione_pos(dati_completo['pos_tags'])
    risultati.append({
        "Corpus": "Totale",
        "Frasi": len(dati_completo['frasi']),
        "Token": len(dati_completo['tokens']),
        "Lunghezza_media_frasi": round(lunghezza_media, 2),
        "Lemmi_distinti": len(set(dati_completo['lemmi'])),
        "Verbi": pos_dist['Verbi'],
        "Aggettivi": pos_dist['Aggettivi'],
        "Nomi": pos_dist['Nomi'],
        "Pronomi": pos_dist['Pronomi'],
        "Entità_nominate": conta_entita_nominate(dati_completo['doc'])
    })

    return pd.DataFrame(risultati)


# ================================
# Analisi e clustering
# ================================
def analizza_tipologia(df, tipologia, n_clusters=6, top_n=5, nome_file_base="output", salva_plot=False):
    """Esegue clustering e visualizzazione UMAP per una specifica tipologia."""
    print(f"\nAnalisi per tipologia: {tipologia}")
    sotto_df = df[df["TIPOLOGIA"] == tipologia].copy().reset_index(drop=True)
    if sotto_df.empty:
        print(f"Nessun dato disponibile per {tipologia}.")
        return

    embeddings = model.encode(sotto_df["CONTENUTO"].tolist(), show_progress_bar=True)
    kmeans = KMeans(n_clusters=n_clusters, random_state=42)
    sotto_df["CLUSTER"] = kmeans.fit_predict(embeddings)

    if n_clusters > 1:
        score = silhouette_score(embeddings, sotto_df["CLUSTER"])
        print(f"Silhouette score ({tipologia}): {score:.3f}")

    # Etichette esemplificative dei cluster
    etichette = {
        0: "Nevrosi e isteria",
        1: "Colpa e trauma",
        2: "Devianza e alienazione",
        3: "Malinconia e dolore",
        4: "Fragilità nervosa e ansia",
        5: "Illusione e speranza"
    }
    sotto_df["ETICHETTA"] = sotto_df["CLUSTER"].map(etichette)

    df_rapp, embeddings_rapp, rappresentative_dict = parole_rappresentative(
        embeddings, sotto_df, cluster_col="CLUSTER", top_n=top_n
    )

    cartella_output = os.path.join("output", nome_file_base)
    os.makedirs(cartella_output, exist_ok=True)

    reducer = umap.UMAP(n_neighbors=10, min_dist=0.1, random_state=42)
    embedding_2d = reducer.fit_transform(embeddings_rapp)

    # Tronca le frasi lunghe per migliorare la leggibilità nel grafico
    def tronca_testo(testo, max_len=50):
        return testo if len(testo) <= max_len else testo[:max_len].rstrip() + "[...]"

    df_rapp["CONTENUTO_TRONCATO"] = df_rapp["CONTENUTO"].apply(tronca_testo)

    # Visualizzazione 2D dei cluster
    plt.figure(figsize=(12, 10))
    plt.scatter(embedding_2d[:, 0], embedding_2d[:, 1], c=df_rapp["CLUSTER"], cmap="tab10", s=60)
    texts = [plt.text(embedding_2d[i, 0], embedding_2d[i, 1], txt, fontsize=9)
             for i, txt in enumerate(df_rapp["CONTENUTO_TRONCATO"])]

    adjust_text(
        texts,
        only_move={'points': 'y', 'texts': 'y'},
        arrowprops=dict(arrowstyle="-", color='gray', alpha=0.5),
        expand_points=(1.2, 1.4),
        expand_text=(1.2, 1.4),
        force_text=0.5,
        lim=100
    )

    # Limiti dinamici per mantenere il testo all’interno del grafico
    x_min, x_max = embedding_2d[:, 0].min(), embedding_2d[:, 0].max()
    y_min, y_max = embedding_2d[:, 1].min(), embedding_2d[:, 1].max()
    plt.xlim(x_min - 1, x_max + 1)
    plt.ylim(y_min - 1, y_max + 1)

    plt.title(f"Cluster {tipologia} - Rappresentanti (Top {top_n})", fontsize=12)
    plt.xlabel("UMAP 1")
    plt.ylabel("UMAP 2")
    plt.tight_layout()

    nome_plot = os.path.join(cartella_output, f"umap_{tipologia.lower()}_{nome_file_base}.png")
    plt.savefig(nome_plot, dpi=300, bbox_inches="tight")
    plt.close()
    print(f"Plot salvato in {nome_plot}")

    nome_excel = os.path.join(cartella_output, f"cluster_{tipologia.lower()}_{nome_file_base}.xlsx")
    salva_excel_auto_width(sotto_df, nome_excel)

    return sotto_df


# ================================
# Esecuzione principale
# ================================
if __name__ == "__main__":
    coppie = [
        {"tabella": "dizionario_leappassionate.xlsx", "corpus": "corpus_leappassionate.txt", "nome_base": "leappassionate"},
        {"tabella": "dizionario_profumo.xlsx", "corpus": "corpus_profumo.txt", "nome_base": "profumo"}
    ]

    os.makedirs("output", exist_ok=True)
    report_list = []
    testi = []
    nomi_corpus = []

    for c in coppie:
        cartella_output = os.path.join("output", c["nome_base"])
        os.makedirs(cartella_output, exist_ok=True)

        testo = leggi_corpora(c["corpus"])
        doc = nlp(testo)
        df = pd.read_excel(c["tabella"])
        testi.append(testo)
        nomi_corpus.append(c["nome_base"].capitalize())

        report_list.append(
            analizza_tipologia(df, "LESSICO", n_clusters=6, top_n=10, nome_file_base=c["nome_base"])
        )
        report_list.append(
            analizza_tipologia(df, "FRASE", n_clusters=6, top_n=5, nome_file_base=c["nome_base"])
        )

    report_finale = pd.concat(report_list, ignore_index=True)
    report_finale["NUM_ELEMENTI_CLUSTER"] = report_finale.groupby("CLUSTER")["CLUSTER"].transform("count")
    report_finale.sort_values(by="NUM_ELEMENTI_CLUSTER", ascending=False, inplace=True)
    report_finale.drop(columns=["NUM_ELEMENTI_CLUSTER"], inplace=True)
    salva_excel_auto_width(report_finale, os.path.join("output", "report_cluster_completo_ordinato.xlsx"))
    print("Report complessivo salvato in output/report_cluster_completo_ordinato.xlsx")

    df_linguistica = analisi_linguistica(testi, nomi_corpus)
    salva_excel_auto_width(df_linguistica, os.path.join("output", "report_analisi_linguistica.xlsx"))
    print("Report di analisi linguistica salvato in output/report_analisi_linguistica.xlsx")
